"""Excel export of Leistungsverzeichnis in Austrian tendering format."""

import io
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.db.models.lv import Leistungsverzeichnis, Leistungsgruppe, Position
from app.db.models.calculation import Berechnungsnachweis
from app.db.models.project import Project


async def export_lv_xlsx(lv_id: UUID, db: AsyncSession) -> bytes:
    """Export an LV as Excel (.xlsx) file.

    Returns bytes of the xlsx file.
    """
    # Load LV with all nested data
    stmt = (
        select(Leistungsverzeichnis)
        .where(Leistungsverzeichnis.id == lv_id)
        .options(
            selectinload(Leistungsverzeichnis.gruppen)
            .selectinload(Leistungsgruppe.positionen)
            .selectinload(Position.berechnungsnachweise),
            selectinload(Leistungsverzeichnis.project),
        )
    )
    result = await db.execute(stmt)
    lv = result.scalars().first()
    if not lv:
        raise ValueError(f"LV {lv_id} not found")

    wb = Workbook()

    # Sheet 1: Zusammenfassung (Summary)
    _create_summary_sheet(wb.active, lv)

    # Sheet 2: Leistungsverzeichnis (Full LV)
    ws_lv = wb.create_sheet("Leistungsverzeichnis")
    _create_lv_sheet(ws_lv, lv)

    # Sheet 3: Berechnungsnachweis (Calculation proof)
    ws_calc = wb.create_sheet("Berechnungsnachweis")
    _create_calculation_sheet(ws_calc, lv)

    # Write to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _create_summary_sheet(ws, lv: Leistungsverzeichnis):
    ws.title = "Zusammenfassung"

    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)

    # Title
    ws["A1"] = "LEISTUNGSVERZEICHNIS"
    ws["A1"].font = header_font
    ws["A3"] = "Projekt:"
    ws["A3"].font = label_font
    ws["B3"] = lv.project.name if lv.project else ""
    ws["A4"] = "Gewerk:"
    ws["A4"].font = label_font
    ws["B4"] = lv.trade
    ws["A5"] = "Status:"
    ws["A5"].font = label_font
    ws["B5"] = lv.status

    # Group summaries
    row = 8
    ws.cell(row=row, column=1, value="LG").font = label_font
    ws.cell(row=row, column=2, value="Bezeichnung").font = label_font
    ws.cell(row=row, column=3, value="Anzahl Positionen").font = label_font
    row += 1

    for gruppe in sorted(lv.gruppen, key=lambda g: g.sort_order):
        ws.cell(row=row, column=1, value=gruppe.nummer)
        ws.cell(row=row, column=2, value=gruppe.bezeichnung)
        ws.cell(row=row, column=3, value=len(gruppe.positionen))
        row += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20


def _create_lv_sheet(ws, lv: Leistungsverzeichnis):
    """Create the main LV sheet with all positions."""
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin"),
    )

    # Headers
    headers = ["Pos.-Nr.", "Kurztext", "Einheit", "Menge", "EP [€]", "GP [€]"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for gruppe in sorted(lv.gruppen, key=lambda g: g.sort_order):
        # Group header
        ws.cell(row=row, column=1, value=gruppe.nummer)
        cell = ws.cell(row=row, column=2, value=gruppe.bezeichnung)
        cell.font = Font(bold=True)
        row += 1

        for pos in sorted(gruppe.positionen, key=lambda p: p.sort_order):
            ws.cell(row=row, column=1, value=pos.positions_nummer)
            ws.cell(row=row, column=2, value=pos.kurztext)
            ws.cell(row=row, column=3, value=pos.einheit)
            menge_cell = ws.cell(row=row, column=4, value=float(pos.menge) if pos.menge else 0)
            menge_cell.number_format = "#,##0.000"
            ep_cell = ws.cell(row=row, column=5, value=float(pos.einheitspreis) if pos.einheitspreis else None)
            ep_cell.number_format = "#,##0.00"
            gp = pos.gesamtpreis
            gp_cell = ws.cell(row=row, column=6, value=gp if gp else None)
            gp_cell.number_format = "#,##0.00"

            # Add langtext in next row if available
            if pos.langtext:
                row += 1
                cell = ws.cell(row=row, column=2, value=pos.langtext)
                cell.font = Font(size=9, italic=True)
                cell.alignment = Alignment(wrap_text=True)

            row += 1

        row += 1  # Empty row between groups

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14


def _create_calculation_sheet(ws, lv: Leistungsverzeichnis):
    """Create the calculation proof sheet (Berechnungsnachweis)."""
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    headers = [
        "Pos.-Nr.", "Raum", "Beschreibung", "Formel",
        "Rohmaß", "Faktor", "Regel", "Abzüge", "Netto", "Einheit"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for gruppe in sorted(lv.gruppen, key=lambda g: g.sort_order):
        for pos in sorted(gruppe.positionen, key=lambda p: p.sort_order):
            for nachweis in pos.berechnungsnachweise:
                ws.cell(row=row, column=1, value=pos.positions_nummer)
                room = nachweis.room
                ws.cell(row=row, column=2, value=room.name if room else "")
                ws.cell(row=row, column=3, value=nachweis.formula_description)
                ws.cell(row=row, column=4, value=nachweis.formula_expression)
                ws.cell(row=row, column=5, value=float(nachweis.raw_quantity)).number_format = "#,##0.000"
                ws.cell(row=row, column=6, value=float(nachweis.onorm_factor)).number_format = "#,##0.0000"
                ws.cell(row=row, column=7, value=nachweis.onorm_paragraph or "")
                # Format deductions
                deductions_str = "; ".join(
                    f"{d.get('opening', '')}: {d.get('area', '')}m² ({'abgezogen' if d.get('deducted') else 'nicht abgezogen'})"
                    for d in (nachweis.deductions or [])
                )
                ws.cell(row=row, column=8, value=deductions_str)
                ws.cell(row=row, column=9, value=float(nachweis.net_quantity)).number_format = "#,##0.000"
                ws.cell(row=row, column=10, value=nachweis.unit)
                row += 1

    # Column widths
    widths = [12, 25, 40, 30, 12, 10, 20, 40, 12, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
